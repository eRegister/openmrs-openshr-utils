# Import openpyxl 
import openpyxl 
from  openpyxl import Workbook
# import load_workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Open the spreadsheet 
def compare_eregister_vs_hie_data(demographics_file, hie_file,facility_name):
    workbook = openpyxl.load_workbook(demographics_file) 
    workbook2 = openpyxl.load_workbook(hie_file) 

    # Get the first sheet 
    sheet = workbook.worksheets[0] 
    sheet2 = workbook2.worksheets[0] 

    # Create a list to store the values 
    
    wb2 = Workbook()
    # set file path
    filepath2="/home/openmrs/openmrs-openshr-utils/data/"+facility_name+"_found_HTS_observations.xlsx"
    # save workbook 
    wb2.save(filepath2)
    wb2 = load_workbook(filepath2)
    
    # select demo.xlsx
    sheet3=wb2.active

    # Create a list to store the values 
    
    wb3 = Workbook()
    # set file path
    filepath3="/home/openmrs/openmrs-openshr-utils/data/"+facility_name+"_missing_HTS_observations_"+datetime.today().strftime('%d_%m_%Y')+".xlsx"
    # save workbook 
    wb3.save(filepath3)
    wb3 = load_workbook(filepath3)
    
    # select demo.xlsx
    sheet4=wb3.active

    # Iterate over the rows in the sheet 
    for row in sheet: 
        count=0
        # Get the value of the first cell 
        # in the row (the "Name" cell) 
        id = row[0].value 
        # name=row[1].value
        found=0
        for row2 in sheet2: 
            count+=1
            id2 = row2[0].value
            # name2 = row2[0].value
            if id[:14]==id2:
                found=1
                # Add the value to the list 
                sheet3.append([id2])
        if found==0:
            sheet4.append([id[:14]]) 
    
    wb2.save(filepath2)
    wb3.save(filepath3)
    
    print("Saved clients with found encounters to '"+filepath2+"'")
    print("Saved clients with missing encounters to '"+filepath3+"'")

    # Print the list of names 
