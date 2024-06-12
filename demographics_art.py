import sys
#pandas library is not visible when running from crontab so appending path site to site packages exposes it
sys.path.append('/home/openmrs/openmrs-openshr-utils/global_variables/')
import openshr_automation_global_variables
sys.path.append(openshr_automation_global_variables.get_server_python_version_dependency_path())
import pandas as pd
import os
import openpyxl
from  openpyxl import Workbook
from datetime import datetime

# import load_workbook
from openpyxl import load_workbook

def demographics(facility_name,sql_exported_file,heina_id):

    #########################################
    ### create a new empty work book
    #########################################
    
    # create Workbook object
    read_file = pd.read_csv (sql_exported_file,header=None)
    read_file.to_csv (sql_exported_file, header=['Identifier', 'givenname', 'familyname', 'address1', 'address2', 'city', 'dateOfBirth', 'phoneNumber', 'identifier', 'gender', 'mothersMaidenName', 'familyname2', 'motherName', 'maritalstatusCode'], index = False)
    
    if os.path.exists('/home/openmrs/openmrs-openshr-utils/converted_ART.xlsx'):
        os.remove('/home/openmrs/openmrs-openshr-utils/converted_ART.xlsx')
    read_file = pd.read_csv (sql_exported_file)
    read_file.to_excel ('/home/openmrs/openmrs-openshr-utils/converted_ART.xlsx', index = False, header=True)
    
    wb2 = Workbook()
    # set file path
    filepath2="/home/openmrs/openmrs-openshr-utils/"+facility_name+"_ARTNew.xlsx"
    # if not os.path.exists("./data/"):
    #     os.mkdir("data")
    if os.path.exists(filepath2):
        os.remove(filepath2)
        
    # save workbook 
    wb2.save(filepath2)
    
    # load demo.xlsx 
    wb2 = load_workbook(filepath2)
    
    # select demo.xlsx
    sheet2=wb2.active
    
    heina = heina_id+'&'+heina_id
    print("heina: "+heina)
    
    # set file path
    filepath = "/home/openmrs/openmrs-openshr-utils/converted_ART.xlsx"
    # load demo.xlsx 
    wb = load_workbook(filename = filepath)
    
    # select demo.xlsx
    sheet=wb.active
    
    row_count = sheet.max_row
    column_count = sheet.max_column
    row_count = sheet.max_row
    column_count = sheet.max_column
    (sheet2.cell(row=1,column=1)).value = "Identifier"
    (sheet2.cell(row=1,column=2)).value = "givenname"
    (sheet2.cell(row=1,column=3)).value = "familyname"
    (sheet2.cell(row=1,column=4)).value = "address1"
    (sheet2.cell(row=1,column=5)).value = "address2"
    (sheet2.cell(row=1,column=6)).value = "city"
    (sheet2.cell(row=1,column=7)).value = "dateOfBirth"
    (sheet2.cell(row=1,column=8)).value = "phoneNumber"
    (sheet2.cell(row=1,column=9)).value = "identifier"
    (sheet2.cell(row=1,column=10)).value = "gender"
    (sheet2.cell(row=1,column=11)).value = "mothersMaidenName"
    (sheet2.cell(row=1,column=12)).value = "familyname2"
    (sheet2.cell(row=1,column=13)).value = "motherName"
    (sheet2.cell(row=1,column=14)).value = "maritalstatusCode"
    
    #iterating through all the columns in the sheet
    for i in range(2,row_count + 1):
    
        identifier = (sheet.cell(row=i,column=1)).value + "&" + heina + "&" + "PI"
        
        givenname = (sheet.cell(row=i,column=2)).value
       # givenname_smallCase = givenname.casefold()
       # givenname_camelCase = givenname_smallCase.capitalize()
        
        familyname = (sheet.cell(row=i,column=3)).value
       # familyname_smallCase = familyname.casefold()
       # familyname_camelCase = familyname_smallCase.capitalize()
        
        address1 = (sheet.cell(row=i,column=4)).value
        address2 = (sheet.cell(row=i,column=5)).value
        city = (sheet.cell(row=i,column=6)).value
        dateOfBirth = (sheet.cell(row=i,column=7)).value
        datetime_object = datetime.strptime(dateOfBirth, '%Y-%M-%d')
        final = datetime_object.strftime("%Y") + datetime_object.strftime("%M") + datetime_object.strftime("%d")
        phoneNuber = (sheet.cell(row=i,column=8)).value
        identifier2 = ""
        gender = (sheet.cell(row=i,column=10)).value
        mothersMaidenName = (sheet.cell(row=i,column=11)).value
        familyname2 = (sheet.cell(row=i,column=12)).value
        motherName = (sheet.cell(row=i,column=13)).value
        maritaltatusCode = (sheet.cell(row=i,column=14)).value
    
    #converting marital statuscods from openrs accordingly    
        if (maritaltatusCode == "2181"):
            maritaltatusCode = "B"
    
        elif(maritaltatusCode == "2182"):
            maritaltatusCode = "M"
            
        elif(maritaltatusCode == "2183"):
            maritaltatusCode = "D"
        
        elif(maritaltatusCode == "2184"):
            maritaltatusCode = "W"
        
        elif(maritaltatusCode == "4173"):
            maritaltatusCode = "S"      
    
        elif(maritaltatusCode == "4178"):
            maritaltatusCode = "A"
        
        elif (maritaltatusCode == "4244"):
            maritaltatusCode = "O"
        
        else:
            maritaltatusCode = "Z"
    
    # writing values from oldsheet to a new sheet      
        (sheet2.cell(row=i,column=1)).value = identifier
        (sheet2.cell(row=i,column=2)).value = givenname
        (sheet2.cell(row=i,column=3)).value = familyname
        (sheet2.cell(row=i,column=4)).value = address1
        (sheet2.cell(row=i,column=5)).value = address2
        (sheet2.cell(row=i,column=6)).value = city
        (sheet2.cell(row=i,column=7)).value = final
        (sheet2.cell(row=i,column=8)).value = phoneNuber
        (sheet2.cell(row=i,column=9)).value = identifier2
        (sheet2.cell(row=i,column=10)).value = gender
        (sheet2.cell(row=i,column=11)).value = mothersMaidenName
        (sheet2.cell(row=i,column=12)).value = familyname2
        (sheet2.cell(row=i,column=13)).value = motherName
        (sheet2.cell(row=i,column=14)).value = maritaltatusCode       
    # save workbook 
    wb2.save(filepath2)
    
    print("Covertion done!!!!")
    return filepath2
